#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""项目计划API - V2.2重构版"""

from fastapi import APIRouter, Depends, HTTPException, Query, UploadFile, File, Form
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import and_, or_, func
from typing import Optional, List
from datetime import datetime, date
from io import BytesIO
from urllib.parse import quote
import os
import uuid

from app.database import get_db
from app.models.wbs_task import WbsTask
from app.models.user import User
from app.models.school import School
from app.models.task_attachment import TaskAttachment
from app.core.security import get_current_user

router = APIRouter()

# 佐证材料上传配置
UPLOAD_SUBDIR = os.path.join("uploads", "task_attachments")
UPLOAD_DIR = os.path.abspath(UPLOAD_SUBDIR)
MAX_FILE_SIZE = 20 * 1024 * 1024  # 20MB
# 类型白名单(扩展名)
ALLOWED_EXTS = {
    ".pdf", ".doc", ".docx", ".xls", ".xlsx", ".ppt", ".pptx",
    ".png", ".jpg", ".jpeg", ".gif", ".bmp", ".webp"
}


@router.get("/project-plan/summary")
def get_project_summary(
    current_user_id: Optional[int] = Query(None, description="当前用户ID"),
    db: Session = Depends(get_db)
):
    """
    获取项目计划汇总数据（4个卡片）
    - 整体进度：按L4完成率计算
    - 进行中任务数
    - 已延期任务数
    - 我的任务数
    """
    # 基础查询：排除已删除
    query = db.query(WbsTask).filter(WbsTask.is_orphan == 0)

    # 总任务数（L4任务）
    l4_tasks = query.filter(WbsTask.work_content_l4 != "").all()
    total_l4 = len(l4_tasks)

    # 已完成的L4任务数
    completed_l4 = len([t for t in l4_tasks if t.status == '已完成'])

    # 整体进度 = 已完成L4 / 总L4
    overall_progress = round((completed_l4 / total_l4 * 100)) if total_l4 > 0 else 0

    # 进行中任务数
    doing_count = query.filter(WbsTask.status == '进行中').count()

    # 已延期任务数（计划结束日期 < 今天 且 状态不是已完成）
    today = date.today()
    delayed_count = query.filter(
        and_(
            WbsTask.plan_end_date < today,
            WbsTask.status != '已完成'
        )
    ).count()

    # 我的任务数
    my_tasks_count = 0
    if current_user_id:
        my_tasks_count = query.filter(
            WbsTask.responsible_person_id == current_user_id
        ).count()

    return {
        "overall_progress": overall_progress,
        "doing_count": doing_count,
        "delayed_count": delayed_count,
        "my_tasks_count": my_tasks_count
    }


@router.get("/project-plan/gantt")
def get_gantt_view(
    status: Optional[str] = Query(None),
    responsible_person_id: Optional[int] = Query(None),
    school_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, description="截止日期起(含),按plan_end_date过滤"),
    date_to: Optional[str] = Query(None, description="截止日期止(含),按plan_end_date过滤"),
    delayed: Optional[bool] = Query(None, description="筛选已延期任务(计划结束日期<今天且状态≠已完成)"),
    db: Session = Depends(get_db)
):
    """
    获取甘特图视图数据
    - 返回L1-L4层级结构
    - 包含时间轴信息
    """
    query = db.query(
        WbsTask,
        User.real_name.label("assignee_name"),
        School.full_name.label("school_name")
    ).outerjoin(
        User, WbsTask.responsible_person_id == User.id
    ).outerjoin(
        School, WbsTask.school_id == School.id
    ).filter(
        WbsTask.is_orphan == 0
    )

    # 筛选条件
    if status:
        query = query.filter(WbsTask.status == status)
    if responsible_person_id:
        query = query.filter(WbsTask.responsible_person_id == responsible_person_id)
    if school_id:
        query = query.filter(WbsTask.school_id == school_id)
    if keyword:
        query = query.filter(
            or_(
                WbsTask.work_content_l4.like(f"%{keyword}%"),
                WbsTask.task_code.like(f"%{keyword}%")
            )
        )
    if date_from:
        query = query.filter(WbsTask.plan_end_date >= date_from)
    if date_to:
        query = query.filter(WbsTask.plan_end_date <= date_to)
    if delayed:
        today = date.today()
        query = query.filter(
            and_(
                WbsTask.plan_end_date < today,
                WbsTask.status != '已完成'
            )
        )

    # 排序：按L1-L4层级排序
    results = query.order_by(
        WbsTask.project_phase_l1,
        WbsTask.sub_phase_l2,
        WbsTask.task_package_l3,
        WbsTask.work_content_l4
    ).all()

    # 组装层级结构
    items = []
    for task, assignee_name, school_name in results:
        items.append({
            "id": task.id,
            "task_code": task.task_code,
            "level": _get_task_level(task),
            "l1": task.project_phase_l1,
            "l2": task.sub_phase_l2,
            "l3": task.task_package_l3,
            "l4": task.work_content_l4,
            "assignee_name": assignee_name or "",
            "school_name": school_name or "",
            "plan_start_date": task.plan_start_date.isoformat() if task.plan_start_date else None,
            "plan_end_date": task.plan_end_date.isoformat() if task.plan_end_date else None,
            "status": task.status,
            "priority": task.priority,
            "stage_type": task.stage_type or "",
            "progress_note": task.progress_note or "",
            "progress": task.progress if task.progress is not None else 0,
            "parent_id": task.parent_id
        })

    return {"items": items}


def _get_task_level(task: WbsTask) -> int:
    """判断任务层级(表无parent_id,按最深非空层级字段推算)"""
    if task.work_detail_l5:
        return 5
    elif task.work_content_l4:
        return 4
    elif task.task_package_l3:
        return 3
    elif task.sub_phase_l2:
        return 2
    else:
        return 1


def _task_path(task: WbsTask, level: int) -> tuple:
    """取任务从L1到指定层级的路径元组,用于父子关系匹配"""
    fields = [
        task.project_phase_l1 or "",
        task.sub_phase_l2 or "",
        task.task_package_l3 or "",
        task.work_content_l4 or "",
        task.work_detail_l5 or "",
    ]
    return tuple(fields[:level])


def _get_direct_children(task: WbsTask, db: Session) -> List[WbsTask]:
    """取某任务的直接子任务(level+1 且 L1..L(level)路径前缀一致,同建设年份)"""
    level = _get_task_level(task)
    if level >= 5:
        return []
    my_path = _task_path(task, level)
    candidates = db.query(WbsTask).filter(
        WbsTask.is_orphan == 0,
        WbsTask.construction_year == task.construction_year,
        WbsTask.id != task.id
    ).all()
    children = []
    for c in candidates:
        if _get_task_level(c) == level + 1 and _task_path(c, level) == my_path:
            children.append(c)
    return children


def _get_ancestors(task: WbsTask, db: Session) -> List[WbsTask]:
    """取某任务的所有祖先任务(由近到远),按层级路径前缀匹配"""
    level = _get_task_level(task)
    ancestors = []
    for anc_level in range(level - 1, 0, -1):
        anc_path = _task_path(task, anc_level)
        candidates = db.query(WbsTask).filter(
            WbsTask.is_orphan == 0,
            WbsTask.construction_year == task.construction_year,
            WbsTask.id != task.id
        ).all()
        for a in candidates:
            if _get_task_level(a) == anc_level and _task_path(a, anc_level) == anc_path:
                ancestors.append(a)
                break
    return ancestors


def _cascade_status_up(task: WbsTask, db: Session):
    """
    父子状态联动(设计§6.7.1),自底向上逐级传导:
    - 规则2: 子任务变"进行中" → 祖先若"待开始"自动变"进行中"
    - 规则3: 父任务全部直接子任务"已完成" → 父任务自动"已完成"
    - 规则4: 父任务全部直接子任务"待开始" → 父任务"待开始"
    """
    for ancestor in _get_ancestors(task, db):
        children = _get_direct_children(ancestor, db)
        if not children:
            continue
        child_statuses = [c.status for c in children]
        if all(s == '已完成' for s in child_statuses):
            ancestor.status = '已完成'
        elif all(s == '待开始' for s in child_statuses):
            ancestor.status = '待开始'
        elif any(s in ('进行中', '已完成', '待补材料', '已延期') for s in child_statuses):
            if ancestor.status == '待开始':
                ancestor.status = '进行中'


@router.get("/project-plan/list")
def get_list_view(
    status: Optional[str] = Query(None),
    responsible_person_id: Optional[int] = Query(None),
    school_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, description="截止日期起(含),按plan_end_date过滤"),
    date_to: Optional[str] = Query(None, description="截止日期止(含),按plan_end_date过滤"),
    delayed: Optional[bool] = Query(None, description="筛选已延期任务(计划结束日期<今天且状态≠已完成)"),
    page: int = Query(1, ge=1),
    page_size: int = Query(50, ge=1, le=200),
    db: Session = Depends(get_db)
):
    """
    获取列表视图数据
    - 层级列表展示
    - 支持分页
    """
    query = db.query(
        WbsTask,
        User.real_name.label("assignee_name"),
        School.full_name.label("school_name")
    ).outerjoin(
        User, WbsTask.responsible_person_id == User.id
    ).outerjoin(
        School, WbsTask.school_id == School.id
    ).filter(
        WbsTask.is_orphan == 0
    )

    # 筛选条件
    if status:
        query = query.filter(WbsTask.status == status)
    if responsible_person_id:
        query = query.filter(WbsTask.responsible_person_id == responsible_person_id)
    if school_id:
        query = query.filter(WbsTask.school_id == school_id)
    if keyword:
        query = query.filter(
            or_(
                WbsTask.work_content_l4.like(f"%{keyword}%"),
                WbsTask.task_code.like(f"%{keyword}%")
            )
        )
    if date_from:
        query = query.filter(WbsTask.plan_end_date >= date_from)
    if date_to:
        query = query.filter(WbsTask.plan_end_date <= date_to)
    if delayed:
        today = date.today()
        query = query.filter(
            and_(
                WbsTask.plan_end_date < today,
                WbsTask.status != '已完成'
            )
        )

    # 总数
    total = query.count()

    # 分页
    offset = (page - 1) * page_size
    results = query.order_by(
        WbsTask.project_phase_l1,
        WbsTask.sub_phase_l2,
        WbsTask.task_package_l3,
        WbsTask.work_content_l4
    ).offset(offset).limit(page_size).all()

    # 组装
    items = []
    for task, assignee_name, school_name in results:
        items.append({
            "id": task.id,
            "task_code": task.task_code,
            "level": _get_task_level(task),
            "l1": task.project_phase_l1,
            "l2": task.sub_phase_l2,
            "l3": task.task_package_l3,
            "l4": task.work_content_l4,
            "assignee_name": assignee_name or "",
            "assignee_id": task.responsible_person_id,
            "school_name": school_name or "",
            "school_id": task.school_id,
            "plan_start_date": task.plan_start_date.isoformat() if task.plan_start_date else None,
            "plan_end_date": task.plan_end_date.isoformat() if task.plan_end_date else None,
            "actual_start_date": task.actual_start_date.isoformat() if task.actual_start_date else None,
            "actual_end_date": task.actual_end_date.isoformat() if task.actual_end_date else None,
            "status": task.status,
            "priority": task.priority,
            "stage_type": task.stage_type or "",
            "progress_note": task.progress_note or "",
            "progress": task.progress if task.progress is not None else 0,
            "parent_id": task.parent_id,
            "deliverables": task.deliverables or ""
        })

    return {"total": total, "items": items}


@router.get("/project-plan/kanban")
def get_kanban_view(
    responsible_person_id: Optional[int] = Query(None),
    school_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None, description="截止日期起(含),按plan_end_date过滤"),
    date_to: Optional[str] = Query(None, description="截止日期止(含),按plan_end_date过滤"),
    db: Session = Depends(get_db)
):
    """
    获取看板视图数据
    - 按状态分列
    - 返回每列的任务卡片
    """
    query = db.query(
        WbsTask,
        User.real_name.label("assignee_name"),
        School.full_name.label("school_name")
    ).outerjoin(
        User, WbsTask.responsible_person_id == User.id
    ).outerjoin(
        School, WbsTask.school_id == School.id
    ).filter(
        WbsTask.is_orphan == 0
    )

    # 筛选条件
    if responsible_person_id:
        query = query.filter(WbsTask.responsible_person_id == responsible_person_id)
    if school_id:
        query = query.filter(WbsTask.school_id == school_id)
    if keyword:
        query = query.filter(
            or_(
                WbsTask.work_content_l4.like(f"%{keyword}%"),
                WbsTask.task_code.like(f"%{keyword}%")
            )
        )
    if date_from:
        query = query.filter(WbsTask.plan_end_date >= date_from)
    if date_to:
        query = query.filter(WbsTask.plan_end_date <= date_to)

    results = query.all()

    # 按状态分组
    kanban_data = {
        "待开始": [],
        "进行中": [],
        "已完成": [],
        "已延期": [],
        "待补材料": []
    }

    today = date.today()

    for task, assignee_name, school_name in results:
        # 计算是否延期（计划结束日期 < 今天 且 未完成）
        is_delayed = (
            task.plan_end_date and
            task.plan_end_date < today and
            task.status != '已完成'
        )

        # 延期任务强制归入"已延期"列
        status_key = '已延期' if is_delayed else task.status

        card = {
            "id": task.id,
            "title": task.work_content_l4 or task.task_package_l3 or task.sub_phase_l2 or task.project_phase_l1,
            "parent": f"{task.project_phase_l1} / {task.sub_phase_l2}" if task.sub_phase_l2 else task.project_phase_l1,
            "assignee_name": assignee_name or "",
            "school_name": school_name or "",
            "is_delayed": is_delayed,
            "delay_days": (today - task.plan_end_date).days if is_delayed else 0
        }

        kanban_data[status_key].append(card)

    # 统计各状态数量
    return {
        "columns": {
            "待开始": {"count": len(kanban_data["待开始"]), "items": kanban_data["待开始"]},
            "进行中": {"count": len(kanban_data["进行中"]), "items": kanban_data["进行中"]},
            "已完成": {"count": len(kanban_data["已完成"]), "items": kanban_data["已完成"]},
            "已延期": {"count": len(kanban_data["已延期"]), "items": kanban_data["已延期"]},
            "待补材料": {"count": len(kanban_data["待补材料"]), "items": kanban_data["待补材料"]}
        }
    }


@router.get("/project-plan/tasks/{task_id}")
def get_task_detail(
    task_id: int,
    db: Session = Depends(get_db)
):
    """获取任务详情"""
    result = db.query(
        WbsTask,
        User.real_name.label("assignee_name"),
        School.full_name.label("school_name")
    ).outerjoin(
        User, WbsTask.responsible_person_id == User.id
    ).outerjoin(
        School, WbsTask.school_id == School.id
    ).filter(
        WbsTask.id == task_id,
        WbsTask.is_orphan == 0
    ).first()

    if not result:
        raise HTTPException(status_code=404, detail="任务不存在")

    task, assignee_name, school_name = result

    return {
        "id": task.id,
        "task_code": task.task_code,
        "construction_year": task.construction_year,
        "project_phase_l1": task.project_phase_l1,
        "sub_phase_l2": task.sub_phase_l2,
        "task_package_l3": task.task_package_l3,
        "work_content_l4": task.work_content_l4,
        "work_detail_l5": task.work_detail_l5,
        "priority": task.priority,
        "stage_type": task.stage_type or "",
        "status": task.status,
        "plan_start_date": task.plan_start_date.isoformat() if task.plan_start_date else None,
        "plan_end_date": task.plan_end_date.isoformat() if task.plan_end_date else None,
        "actual_start_date": task.actual_start_date.isoformat() if task.actual_start_date else None,
        "actual_end_date": task.actual_end_date.isoformat() if task.actual_end_date else None,
        "responsible_person_id": task.responsible_person_id,
        "assignee_name": assignee_name or "",
        "school_id": task.school_id,
        "school_name": school_name or "",
        "progress_note": task.progress_note or "",
        "progress": task.progress if task.progress is not None else 0,
        "parent_id": task.parent_id,
        "deliverables": task.deliverables or "",
        "created_at": task.created_at.isoformat() if task.created_at else None,
        "updated_at": task.updated_at.isoformat() if task.updated_at else None
    }


@router.post("/project-plan/tasks", status_code=201)
def create_task(
    task_data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    新增任务（方案A：基于 parent_id 挂靠）
    - task_code 后台自动生成（无需前端传）
    - 父任务决定层级：子层级 = 父层级 + 1，并继承父的 L1-L(父层级) 路径
    - L1 任务 parent_id 为空
    """
    task_name = (task_data.get("task_name") or "").strip()
    if not task_name:
        raise HTTPException(status_code=400, detail="任务名称不能为空")

    # 日期倒挂拦截(页面/接口统一在此把关)
    _assert_date_order(task_data.get("plan_start_date"), task_data.get("plan_end_date"))

    parent_id = task_data.get("parent_id")
    parent = None
    if parent_id:
        parent = db.query(WbsTask).filter(
            WbsTask.id == parent_id, WbsTask.is_orphan == 0
        ).first()
        if not parent:
            raise HTTPException(status_code=400, detail="父任务不存在")

    # 计算层级与 L1-L5 路径字段
    parent_level = _get_task_level(parent) if parent else 0
    new_level = parent_level + 1
    if new_level > 5:
        raise HTTPException(status_code=400, detail="任务层级最多 5 级，该父任务已是最末级")

    # 继承父路径，本级填 task_name
    fields = {
        "project_phase_l1": parent.project_phase_l1 if parent else "",
        "sub_phase_l2": parent.sub_phase_l2 if parent else "",
        "task_package_l3": parent.task_package_l3 if parent else "",
        "work_content_l4": parent.work_content_l4 if parent else "",
        "work_detail_l5": parent.work_detail_l5 if parent else "",
    }
    level_field = {1: "project_phase_l1", 2: "sub_phase_l2", 3: "task_package_l3",
                   4: "work_content_l4", 5: "work_detail_l5"}[new_level]
    fields[level_field] = task_name

    # 父子范围校验:子任务须完全落在父范围内
    # 未指定策略 → 409 让页面弹窗二选一;指定了 → 同事务内按策略调整
    eff_start = task_data.get("plan_start_date")
    eff_end = task_data.get("plan_end_date")
    conflict_notes = []
    if parent:
        violation = _range_violation(parent, eff_start, eff_end)
        if violation:
            strategy = task_data.get("on_date_conflict")
            if not strategy:
                _raise_range_conflict(violation)
            eff_start, eff_end, conflict_notes = _resolve_date_conflict(
                parent, eff_start, eff_end, strategy, db
            )

    # 建设年份：继承父任务，否则用前端传值或当前年
    construction_year = (parent.construction_year if parent
                         else task_data.get("construction_year") or str(date.today().year))

    # 自动生成唯一 task_code：WBS-AUTO-{最大序号+1}
    auto_code = _gen_task_code(db)

    new_task = WbsTask(
        task_code=auto_code,
        construction_year=construction_year,
        project_phase_l1=fields["project_phase_l1"],
        sub_phase_l2=fields["sub_phase_l2"],
        task_package_l3=fields["task_package_l3"],
        work_content_l4=fields["work_content_l4"],
        work_detail_l5=fields["work_detail_l5"],
        parent_id=parent_id or None,
        priority=task_data.get("priority", "中"),
        stage_type=task_data.get("stage_type"),
        status=task_data.get("status", "待开始"),
        progress=_clamp_progress(task_data.get("progress")),
        plan_start_date=eff_start,
        plan_end_date=eff_end,
        responsible_person_id=task_data.get("responsible_person_id"),
        school_id=task_data.get("school_id"),
        progress_note=task_data.get("progress_note", ""),
        deliverables=task_data.get("deliverables", "")
    )

    db.add(new_task)
    db.commit()
    db.refresh(new_task)

    resp = {"id": new_task.id, "message": "任务创建成功"}
    if conflict_notes:
        resp["date_adjustments"] = conflict_notes
        resp["message"] = "任务创建成功,已按所选方案调整计划日期"
    return resp


def _clamp_progress(val) -> int:
    """进度归一到 0-100 整数，非法值归 0"""
    try:
        return max(0, min(100, int(val)))
    except (TypeError, ValueError):
        return 0


def _to_date(val):
    """把入参(字符串/date/datetime)统一转 date;空值或格式非法返回 None"""
    if val is None or val == "":
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    s = str(val).strip()
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _assert_date_order(start, end):
    """校验计划结束不早于计划开始,违反则 400(页面/接口统一入口)"""
    s, e = _to_date(start), _to_date(end)
    if s and e and e < s:
        raise HTTPException(
            status_code=400,
            detail=f"计划结束({e.isoformat()})不能早于计划开始({s.isoformat()})"
        )


# ============ 父子日期范围校验(子任务须完全落在父范围内) ============

def _ancestor_chain(task: WbsTask, db: Session) -> List[WbsTask]:
    """
    按 parent_id 向上取祖先链(由近到远)。
    与既有 _get_ancestors(按 L1-L5 路径前缀匹配)不同:此处走 parent_id,
    与前端树形/导入建链口径一致(CLAUDE.md: 层级树形用 parent_id,不用路径硬凑)。
    带环保护,避免脏数据造成死循环。
    """
    chain = []
    seen = {task.id}
    cur = task
    while cur.parent_id:
        parent = db.query(WbsTask).filter(
            WbsTask.id == cur.parent_id, WbsTask.is_orphan == 0
        ).first()
        if not parent or parent.id in seen:
            break
        chain.append(parent)
        seen.add(parent.id)
        cur = parent
    return chain


def _range_violation(parent: WbsTask, start, end):
    """
    判断子任务 [start,end] 是否越出父范围。
    判定口径(用户确认): 开始 >= 父开始 且 结束 <= 父结束,否则越界。
    返回 None 表示合法,否则返回越界描述 dict。
    """
    if parent is None:
        return None
    ps, pe = _to_date(start), _to_date(end)
    fs, fe = _to_date(parent.plan_start_date), _to_date(parent.plan_end_date)
    if not (ps and pe and fs and fe):
        return None  # 任一端缺失无法比较,交由必填/倒挂校验处理

    early = ps < fs
    late = pe > fe
    if not (early or late):
        return None

    reasons = []
    if early:
        reasons.append(f"开始({ps.isoformat()})早于父任务开始({fs.isoformat()})")
    if late:
        reasons.append(f"结束({pe.isoformat()})晚于父任务结束({fe.isoformat()})")

    return {
        "parent_id": parent.id,
        "parent_name": _display_name(parent),
        "parent_start": fs.isoformat(),
        "parent_end": fe.isoformat(),
        "child_start": ps.isoformat(),
        "child_end": pe.isoformat(),
        "starts_early": early,
        "ends_late": late,
        "message": "子任务" + "、".join(reasons),
        # 方案A: 放大父节点到覆盖子任务
        "expand_parent_to": {
            "start": min(ps, fs).isoformat(),
            "end": max(pe, fe).isoformat()
        },
        # 方案B: 把子任务夹回父范围内
        "clamp_child_to": {
            "start": max(ps, fs).isoformat(),
            "end": min(pe, fe).isoformat()
        }
    }


def _display_name(task: WbsTask) -> str:
    """取任务本级显示名(最深非空层级字段)"""
    return (task.work_detail_l5 or task.work_content_l4 or task.task_package_l3
            or task.sub_phase_l2 or task.project_phase_l1 or f"任务#{task.id}")


VALID_CONFLICT_STRATEGIES = ("expand_parent", "clamp_child")


def _raise_range_conflict(violation: dict):
    """越界且未指定处理策略 → 409,前端据此弹窗给两个选项"""
    raise HTTPException(
        status_code=409,
        detail={
            "code": "DATE_RANGE_CONFLICT",
            "message": violation["message"],
            "violation": violation,
            "options": [
                {"strategy": "expand_parent", "label": "按子任务放大父节点",
                 "preview": violation["expand_parent_to"]},
                {"strategy": "clamp_child", "label": "按父节点规整子任务",
                 "preview": violation["clamp_child_to"]}
            ]
        }
    )


def _expand_ancestors(parent: WbsTask, start, end, db: Session):
    """
    方案A: 放大父节点覆盖 [start,end],并逐级向上放大直到某祖先已能覆盖。
    调用方负责 commit(与主操作同事务)。
    """
    s, e = _to_date(start), _to_date(end)
    touched = []
    node = parent
    while node is not None:
        ns, ne = _to_date(node.plan_start_date), _to_date(node.plan_end_date)
        new_s = min(s, ns) if ns else s
        new_e = max(e, ne) if ne else e
        if (new_s, new_e) == (ns, ne):
            break  # 本级已覆盖,上层必然也覆盖,链条终止
        node.plan_start_date = new_s
        node.plan_end_date = new_e
        touched.append({"id": node.id, "name": _display_name(node),
                        "range": f"{new_s.isoformat()} ~ {new_e.isoformat()}"})
        s, e = new_s, new_e
        node = db.query(WbsTask).filter(
            WbsTask.id == node.parent_id, WbsTask.is_orphan == 0
        ).first() if node.parent_id else None
    return touched


def _resolve_date_conflict(parent: WbsTask, start, end, strategy: str, db: Session):
    """
    按策略处理越界,返回 (最终子任务start, 最终子任务end, 调整说明list)。
    expand_parent: 子任务日期不动,放大父及必要的祖先
    clamp_child:   父不动,子任务日期夹回父范围
    """
    if strategy not in VALID_CONFLICT_STRATEGIES:
        raise HTTPException(
            status_code=400,
            detail=f"on_date_conflict 非法,应为 {'/'.join(VALID_CONFLICT_STRATEGIES)}"
        )
    s, e = _to_date(start), _to_date(end)
    if strategy == "expand_parent":
        touched = _expand_ancestors(parent, s, e, db)
        return s, e, touched

    # clamp_child
    fs, fe = _to_date(parent.plan_start_date), _to_date(parent.plan_end_date)
    new_s, new_e = max(s, fs), min(e, fe)
    if new_e < new_s:
        # 子任务与父范围完全无交集,夹取会得到倒挂区间,只能提示改父或改子
        raise HTTPException(
            status_code=400,
            detail=(f"子任务({s.isoformat()}~{e.isoformat()})与父任务范围"
                    f"({fs.isoformat()}~{fe.isoformat()})无重叠,无法规整;"
                    f"请改用「放大父节点」或手工调整日期")
        )
    note = [{"id": None, "name": "本任务",
             "range": f"{new_s.isoformat()} ~ {new_e.isoformat()}"}]
    return new_s, new_e, note


def _gen_task_code(db: Session) -> str:
    """生成唯一任务编码 WBS-AUTO-{序号}"""
    import re
    rows = db.query(WbsTask.task_code).filter(
        WbsTask.task_code.like("WBS-AUTO-%")
    ).all()
    max_n = 0
    for (code,) in rows:
        m = re.search(r"WBS-AUTO-(\d+)", code or "")
        if m:
            max_n = max(max_n, int(m.group(1)))
    return f"WBS-AUTO-{max_n + 1:04d}"


@router.put("/project-plan/tasks/{task_id}")
def update_task(
    task_id: int,
    task_data: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    更新任务
    - 支持部分字段更新
    - 父子任务联动规则
    """
    task = db.query(WbsTask).filter(
        WbsTask.id == task_id,
        WbsTask.is_orphan == 0
    ).first()

    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    # 日期倒挂拦截:部分更新场景下,未传的那一端取库中现值再比较
    # (行内编辑常只改单个日期字段,只比 payload 内部会漏判)
    date_touched = "plan_start_date" in task_data or "plan_end_date" in task_data
    conflict_notes = []
    if date_touched:
        eff_start = task_data.get("plan_start_date", task.plan_start_date)
        eff_end = task_data.get("plan_end_date", task.plan_end_date)
        _assert_date_order(eff_start, eff_end)

        # 父子范围校验:改后的本任务须仍落在父范围内
        parent = db.query(WbsTask).filter(
            WbsTask.id == task.parent_id, WbsTask.is_orphan == 0
        ).first() if task.parent_id else None
        if parent:
            violation = _range_violation(parent, eff_start, eff_end)
            if violation:
                strategy = task_data.get("on_date_conflict")
                if not strategy:
                    _raise_range_conflict(violation)
                eff_start, eff_end, conflict_notes = _resolve_date_conflict(
                    parent, eff_start, eff_end, strategy, db
                )
                task_data = {**task_data,
                             "plan_start_date": eff_start,
                             "plan_end_date": eff_end}

        # 反向:缩小本任务范围可能把自己的子任务挤出去
        kids = db.query(WbsTask).filter(
            WbsTask.parent_id == task.id, WbsTask.is_orphan == 0
        ).all()
        if kids:
            ns, ne = _to_date(eff_start), _to_date(eff_end)
            outside = []
            for k in kids:
                ks, ke = _to_date(k.plan_start_date), _to_date(k.plan_end_date)
                if ks and ke and ns and ne and (ks < ns or ke > ne):
                    outside.append((k, ks, ke))
            if outside:
                strategy = task_data.get("on_date_conflict")
                if not strategy:
                    names = "、".join(_display_name(k) for k, _, _ in outside[:3])
                    more = f" 等{len(outside)}个" if len(outside) > 3 else ""
                    raise HTTPException(
                        status_code=409,
                        detail={
                            "code": "CHILD_RANGE_CONFLICT",
                            "message": (f"调整后范围装不下子任务({names}{more}),"
                                        f"请选择处理方式"),
                            "affected_children": [
                                {"id": k.id, "name": _display_name(k),
                                 "range": f"{ks.isoformat()} ~ {ke.isoformat()}"}
                                for k, ks, ke in outside
                            ],
                            "options": [
                                {"strategy": "expand_parent",
                                 "label": "保留子任务范围,放大本任务",
                                 "preview": {
                                     "start": min([ns] + [k[1] for k in outside]).isoformat(),
                                     "end": max([ne] + [k[2] for k in outside]).isoformat()
                                 }},
                                {"strategy": "clamp_child",
                                 "label": "按本任务范围规整这些子任务",
                                 "preview": {"start": ns.isoformat(), "end": ne.isoformat()}}
                            ]
                        }
                    )
                if strategy == "expand_parent":
                    # 本任务反向放大到覆盖全部子任务,并向上传导
                    eff_start = min([ns] + [k[1] for k in outside])
                    eff_end = max([ne] + [k[2] for k in outside])
                    task_data = {**task_data,
                                 "plan_start_date": eff_start,
                                 "plan_end_date": eff_end}
                    conflict_notes.append({
                        "id": task.id, "name": _display_name(task),
                        "range": f"{eff_start.isoformat()} ~ {eff_end.isoformat()}"
                    })
                    up_parent = db.query(WbsTask).filter(
                        WbsTask.id == task.parent_id, WbsTask.is_orphan == 0
                    ).first() if task.parent_id else None
                    if up_parent:
                        conflict_notes += _expand_ancestors(up_parent, eff_start, eff_end, db)
                elif strategy == "clamp_child":
                    for k, ks, ke in outside:
                        k.plan_start_date = max(ks, ns)
                        k.plan_end_date = min(ke, ne)
                        if k.plan_end_date < k.plan_start_date:
                            raise HTTPException(
                                status_code=400,
                                detail=(f"子任务「{_display_name(k)}」"
                                        f"({ks.isoformat()}~{ke.isoformat()})"
                                        f"与新范围无重叠,无法规整")
                            )
                        conflict_notes.append({
                            "id": k.id, "name": _display_name(k),
                            "range": (f"{k.plan_start_date.isoformat()} ~ "
                                      f"{k.plan_end_date.isoformat()}")
                        })

    # 规则5: 父任务不可越级完成——有未完成直接子任务时禁止手动改为"已完成"
    new_status = task_data.get("status")
    if new_status == '已完成' and task.status != '已完成':
        children = _get_direct_children(task, db)
        unfinished = [c for c in children if c.status != '已完成']
        if unfinished:
            raise HTTPException(
                status_code=400,
                detail=f"该任务还有 {len(unfinished)} 个子任务未完成,不能直接标记为已完成"
            )

    status_changed = new_status is not None and new_status != task.status

    # 更新字段(on_date_conflict 是控制参数,非表字段,跳过)
    for key, value in task_data.items():
        if key == "on_date_conflict":
            continue
        if hasattr(task, key) and value is not None:
            setattr(task, key, value)

    # 状态变更 → 触发父子联动向上传导(规则2/3/4)
    if status_changed:
        _cascade_status_up(task, db)

    db.commit()
    db.refresh(task)

    resp = {"message": "任务更新成功", "id": task.id}
    if conflict_notes:
        resp["date_adjustments"] = conflict_notes
        resp["message"] = "任务更新成功,已按所选方案调整关联日期"
    return resp


@router.delete("/project-plan/tasks/{task_id}", status_code=204)
def delete_task(
    task_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    删除任务（软删除）
    - 标记 is_orphan = 1
    """
    task = db.query(WbsTask).filter(
        WbsTask.id == task_id,
        WbsTask.is_orphan == 0
    ).first()

    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    task.is_orphan = 1
    db.commit()

    return None


@router.post("/project-plan/tasks/check-date-range")
def check_date_range(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    预检:子任务日期是否越出父任务范围(不落库)。
    payload: { parent_id, plan_start_date, plan_end_date, task_id? }
      - task_id 可选,编辑已有任务时传,用于计算「规整子任务」是否会挤压其自身子任务
    返回: { ok: bool, violation: {...}|None, expand_chain: [...], clamp_impact: {...} }
      - expand_chain: 选方案A时会被连带放大的祖先链(父装不下时继续向上)
      - clamp_impact: 选方案B时会越界的自身子任务(供前端提示)
    """
    _assert_date_order(payload.get("plan_start_date"), payload.get("plan_end_date"))

    parent_id = payload.get("parent_id")
    if not parent_id:
        return {"ok": True, "violation": None, "expand_chain": [], "clamp_impact": None}

    parent = db.query(WbsTask).filter(
        WbsTask.id == parent_id, WbsTask.is_orphan == 0
    ).first()
    if not parent:
        raise HTTPException(status_code=400, detail="父任务不存在")

    v = _range_violation(parent, payload.get("plan_start_date"), payload.get("plan_end_date"))
    if v is None:
        return {"ok": True, "violation": None, "expand_chain": [], "clamp_impact": None}

    # 方案A预览:逐级向上,父放大后若仍越出祖父范围则继续放大
    expand_chain = []
    cur_start = _to_date(v["expand_parent_to"]["start"])
    cur_end = _to_date(v["expand_parent_to"]["end"])
    node = parent
    while node is not None:
        expand_chain.append({
            "id": node.id,
            "name": _display_name(node),
            "from": f"{_to_date(node.plan_start_date)} ~ {_to_date(node.plan_end_date)}",
            "to": f"{cur_start.isoformat()} ~ {cur_end.isoformat()}"
        })
        grand = db.query(WbsTask).filter(
            WbsTask.id == node.parent_id, WbsTask.is_orphan == 0
        ).first() if node.parent_id else None
        if grand is None:
            break
        gs, ge = _to_date(grand.plan_start_date), _to_date(grand.plan_end_date)
        if gs is None or ge is None:
            break
        if cur_start >= gs and cur_end <= ge:
            break  # 祖父已能覆盖,链条终止
        cur_start = min(cur_start, gs)
        cur_end = max(cur_end, ge)
        node = grand

    # 方案B预览:夹回父范围后,自身既有子任务是否被挤出
    clamp_impact = None
    task_id = payload.get("task_id")
    if task_id:
        cs = _to_date(v["clamp_child_to"]["start"])
        ce = _to_date(v["clamp_child_to"]["end"])
        kids = db.query(WbsTask).filter(
            WbsTask.parent_id == task_id, WbsTask.is_orphan == 0
        ).all()
        outside = []
        for k in kids:
            ks, ke = _to_date(k.plan_start_date), _to_date(k.plan_end_date)
            if ks and ke and (ks < cs or ke > ce):
                outside.append({"id": k.id, "name": _display_name(k),
                                "range": f"{ks.isoformat()} ~ {ke.isoformat()}"})
        if outside:
            clamp_impact = {"outside_children": outside, "count": len(outside)}

    return {"ok": False, "violation": v, "expand_chain": expand_chain, "clamp_impact": clamp_impact}


@router.post("/project-plan/tasks/batch")
def batch_update_tasks(
    payload: dict,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    批量操作(设计§6.4)
    - action=status: 批量更新状态(含越级完成拦截+父子联动)
    - action=assignee: 批量分配责任人
    - action=delete: 批量软删除
    payload: { task_ids: [int], action: str, value: any }
    """
    task_ids = payload.get("task_ids") or []
    action = payload.get("action")
    value = payload.get("value")

    if not task_ids:
        raise HTTPException(status_code=400, detail="未选择任务")
    if action not in ("status", "assignee", "delete"):
        raise HTTPException(status_code=400, detail="不支持的操作类型")

    tasks = db.query(WbsTask).filter(
        WbsTask.id.in_(task_ids),
        WbsTask.is_orphan == 0
    ).all()

    affected = 0
    skipped = []
    for task in tasks:
        if action == "delete":
            task.is_orphan = 1
            affected += 1
        elif action == "assignee":
            task.responsible_person_id = value
            affected += 1
        elif action == "status":
            # 越级完成拦截(规则5)
            if value == '已完成' and task.status != '已完成':
                children = _get_direct_children(task, db)
                if [c for c in children if c.status != '已完成']:
                    skipped.append(task.task_code)
                    continue
            changed = value != task.status
            task.status = value
            if changed:
                _cascade_status_up(task, db)
            affected += 1

    db.commit()
    result = {"message": "批量操作完成", "affected": affected}
    if skipped:
        result["skipped"] = skipped
        result["message"] = f"完成{affected}项,{len(skipped)}项因有未完成子任务被跳过"
    return result


@router.get("/project-plan/filters")
def get_filter_options(db: Session = Depends(get_db)):
    """
    获取筛选选项
    - 状态列表
    - 责任人列表
    - 学校列表
    """
    # 状态列表
    statuses = ['待开始', '进行中', '已完成', '已延期', '待补材料']

    # 责任人列表
    assignees = db.query(
        User.id,
        User.real_name
    ).filter(
        User.status == "启用"
    ).all()

    # 学校列表
    schools = db.query(
        School.id,
        School.full_name
    ).all()

    return {
        "statuses": statuses,
        "assignees": [{"id": a.id, "name": a.real_name} for a in assignees],
        "schools": [{"id": s.id, "name": s.full_name} for s in schools]
    }


@router.get("/project-plan/parent-options")
def get_parent_options(
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    可选父任务列表（新增任务表单下拉用）
    - 返回 L1-L4 任务（L5 已是最末级，不能作父）
    - label 带完整路径 + 层级，方便识别
    """
    tasks = db.query(WbsTask).filter(WbsTask.is_orphan == 0).order_by(
        WbsTask.project_phase_l1, WbsTask.sub_phase_l2,
        WbsTask.task_package_l3, WbsTask.work_content_l4
    ).all()

    options = []
    for t in tasks:
        level = _get_task_level(t)
        if level >= 5:
            continue  # L5 不能有子任务
        # 路径：本级及以上非空字段用 / 连接
        path_parts = [p for p in [
            t.project_phase_l1, t.sub_phase_l2, t.task_package_l3, t.work_content_l4
        ][:level] if p]
        options.append({
            "id": t.id,
            "level": level,
            "label": f"{' / '.join(path_parts)} (L{level})",
            # 供前端在选定父任务后把子任务日历限制在父范围内
            "plan_start_date": t.plan_start_date.isoformat() if t.plan_start_date else None,
            "plan_end_date": t.plan_end_date.isoformat() if t.plan_end_date else None
        })

    return {"items": options}


# ========== 佐证材料上传(设计§6.7.2 待补材料流转) ==========

@router.get("/project-plan/tasks/{task_id}/attachments")
def list_attachments(task_id: int, db: Session = Depends(get_db)):
    """获取任务的佐证材料列表"""
    task = db.query(WbsTask).filter(WbsTask.id == task_id, WbsTask.is_orphan == 0).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    rows = db.query(TaskAttachment).filter(
        TaskAttachment.task_id == task_id
    ).order_by(TaskAttachment.uploaded_at.desc()).all()

    return {
        "items": [
            {
                "id": a.id,
                "file_name": a.file_name,
                "file_url": f"/{a.file_path.replace(os.sep, '/')}",
                "file_size": a.file_size,
                "description": a.description or "",
                "uploaded_by": a.uploaded_by,
                "uploaded_at": a.uploaded_at.isoformat() if a.uploaded_at else None
            }
            for a in rows
        ]
    }


@router.post("/project-plan/tasks/{task_id}/attachments", status_code=201)
async def upload_attachment(
    task_id: int,
    file: UploadFile = File(...),
    description: Optional[str] = Form(None),
    uploaded_by: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    上传佐证材料
    - 类型白名单 + 20MB 大小限制 + UUID 重命名(防路径穿越/覆盖)
    """
    task = db.query(WbsTask).filter(WbsTask.id == task_id, WbsTask.is_orphan == 0).first()
    if not task:
        raise HTTPException(status_code=404, detail="任务不存在")

    # 类型白名单校验(按扩展名)
    orig_name = file.filename or "unnamed"
    ext = os.path.splitext(orig_name)[1].lower()
    if ext not in ALLOWED_EXTS:
        raise HTTPException(
            status_code=400,
            detail=f"不支持的文件类型 {ext},仅允许 PDF/Office文档/图片"
        )

    # 读取内容并校验大小
    content = await file.read()
    if len(content) > MAX_FILE_SIZE:
        raise HTTPException(status_code=400, detail="文件超过 20MB 限制")
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="文件为空")

    # UUID 重命名,只保留扩展名 → 杜绝原始文件名带来的路径穿越/覆盖
    os.makedirs(UPLOAD_DIR, exist_ok=True)
    stored_name = f"{uuid.uuid4().hex}{ext}"
    abs_path = os.path.join(UPLOAD_DIR, stored_name)
    with open(abs_path, "wb") as f:
        f.write(content)

    # 相对路径入库(供 StaticFiles 提供访问)
    rel_path = os.path.join(UPLOAD_SUBDIR, stored_name)
    attachment = TaskAttachment(
        task_id=task_id,
        file_name=orig_name,
        file_path=rel_path,
        file_size=len(content),
        description=description,
        uploaded_by=uploaded_by
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)

    return {
        "id": attachment.id,
        "file_name": attachment.file_name,
        "file_url": f"/{rel_path.replace(os.sep, '/')}",
        "message": "上传成功"
    }


@router.delete("/project-plan/attachments/{attachment_id}", status_code=204)
def delete_attachment(
    attachment_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """删除佐证材料(同时删除物理文件)"""
    attachment = db.query(TaskAttachment).filter(
        TaskAttachment.id == attachment_id
    ).first()
    if not attachment:
        raise HTTPException(status_code=404, detail="材料不存在")

    # 删除物理文件(不存在则忽略)
    abs_path = os.path.abspath(attachment.file_path)
    if os.path.isfile(abs_path):
        try:
            os.remove(abs_path)
        except OSError:
            pass

    db.delete(attachment)
    db.commit()
    return None


# ============ 导入/导出 ============

# 导出/导入共用列结构(顺序即列顺序),形成"导出→改→导入"闭环
EXPORT_COLUMNS = [
    "L1阶段", "L2子阶段", "L3任务包", "任务名称(L4)", "责任人", "学校",
    "计划开始", "计划结束", "状态", "优先级", "进度%", "进展说明"
]
# 权威枚举(依据 ddl.sql / CLAUDE.md)
_VALID_STATUS = {"待开始", "进行中", "已完成", "已延期", "待补材料"}
_VALID_PRIORITY = {"高", "中", "低"}


def _query_tasks_for_export(db: Session, status, responsible_person_id,
                            school_id, keyword, date_from, date_to, delayed):
    """按筛选条件查询任务(与 list 视图口径一致,但不分页)"""
    query = db.query(
        WbsTask,
        User.real_name.label("assignee_name"),
        School.full_name.label("school_name")
    ).outerjoin(
        User, WbsTask.responsible_person_id == User.id
    ).outerjoin(
        School, WbsTask.school_id == School.id
    ).filter(
        WbsTask.is_orphan == 0,
        # 只导末级实际任务(L4/L5);L1/L2/L3 纯结构节点导入时会自动重建,不导出
        WbsTask.work_content_l4 != ""
    )

    if status:
        query = query.filter(WbsTask.status == status)
    if responsible_person_id:
        query = query.filter(WbsTask.responsible_person_id == responsible_person_id)
    if school_id:
        query = query.filter(WbsTask.school_id == school_id)
    if keyword:
        query = query.filter(or_(
            WbsTask.work_content_l4.like(f"%{keyword}%"),
            WbsTask.task_code.like(f"%{keyword}%")
        ))
    if date_from:
        query = query.filter(WbsTask.plan_end_date >= date_from)
    if date_to:
        query = query.filter(WbsTask.plan_end_date <= date_to)
    if delayed:
        today = date.today()
        query = query.filter(and_(
            WbsTask.plan_end_date < today,
            WbsTask.status != '已完成'
        ))

    return query.order_by(
        WbsTask.project_phase_l1,
        WbsTask.sub_phase_l2,
        WbsTask.task_package_l3,
        WbsTask.work_content_l4
    ).all()


@router.get("/project-plan/export")
def export_tasks(
    status: Optional[str] = Query(None),
    responsible_person_id: Optional[int] = Query(None),
    school_id: Optional[int] = Query(None),
    keyword: Optional[str] = Query(None),
    date_from: Optional[str] = Query(None),
    date_to: Optional[str] = Query(None),
    delayed: Optional[bool] = Query(None),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    导出项目计划为 xlsx(应用当前筛选,不分页)
    - 列结构与导入模板一致,便于"导出→修改→导入"闭环
    - 整份计划批量下发,须登录后调用(与 import 一致)
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    results = _query_tasks_for_export(
        db, status, responsible_person_id, school_id,
        keyword, date_from, date_to, delayed
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "项目计划"

    # 表头样式
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1677FF")
    center = Alignment(horizontal="center", vertical="center")
    ws.append(EXPORT_COLUMNS)
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center

    # 数据行
    for task, assignee_name, school_name in results:
        ws.append([
            task.project_phase_l1 or "",
            task.sub_phase_l2 or "",
            task.task_package_l3 or "",
            task.work_content_l4 or "",
            assignee_name or "",
            school_name or "",
            task.plan_start_date.isoformat() if task.plan_start_date else "",
            task.plan_end_date.isoformat() if task.plan_end_date else "",
            task.status or "",
            task.priority or "",
            task.progress if task.progress is not None else 0,
            task.progress_note or ""
        ])

    # 列宽
    widths = [16, 16, 20, 28, 10, 16, 12, 12, 10, 8, 8, 30]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[chr(64 + i)].width = w
    ws.freeze_panes = "A2"

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"项目计划_{date.today().strftime('%Y%m%d')}.xlsx"
    # RFC 5987 编码中文文件名,避免非 ASCII 报错
    disposition = f"attachment; filename*=UTF-8''{quote(filename)}"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": disposition}
    )


def _cell_str(v) -> str:
    """单元格值转干净字符串(处理 None / 数字 / 日期)"""
    if v is None:
        return ""
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return str(v).strip()


def _parse_date_cell(v):
    """解析日期单元格,失败返回 None(与页面入口共用 _to_date 的解析规则)"""
    return _to_date(v)


def _find_or_create_structure(db, level, path, year, cache, seed_start, seed_end):
    """
    find-or-create 结构节点(仅 L1/L2,与现有数据约定一致:结构只到 L2)。
    path: (l1,) / (l1, l2);level=1 建 L1 节点,level=2 建 L2 节点(L1不存在则递归建)。
    结构节点特征:本级及以上字段有值、更深字段全空、work_content_l4 为空。
    cache: 本次导入内 (level, path)->WbsTask 缓存;seed_*: 新建时的占位日期。
    """
    key = (level, path)
    if key in cache:
        return cache[key]

    # 深层"应为空"的字段可能是空串或 NULL(历史数据混用),两者都要匹配,
    # 否则 SQL 中 NULL = '' 恒为 false,会漏掉现有结构节点导致重复新建。
    def _empty(col):
        return (col == "") | (col.is_(None))

    q = db.query(WbsTask).filter(WbsTask.is_orphan == 0)
    q = q.filter(WbsTask.project_phase_l1 == path[0])
    if level >= 2:
        q = q.filter(WbsTask.sub_phase_l2 == path[1])
    else:
        q = q.filter(_empty(WbsTask.sub_phase_l2))
    q = q.filter(_empty(WbsTask.task_package_l3))
    q = q.filter(_empty(WbsTask.work_content_l4))
    q = q.filter(_empty(WbsTask.work_detail_l5))
    node = q.first()
    if node:
        cache[key] = node
        return node

    parent = None
    if level >= 2:
        parent = _find_or_create_structure(db, 1, path[:1], year, cache, seed_start, seed_end)

    node = WbsTask(
        task_code=_gen_task_code(db),
        construction_year=year,
        project_phase_l1=path[0],
        sub_phase_l2=path[1] if level >= 2 else "",
        task_package_l3="",
        work_content_l4="",
        work_detail_l5="",
        parent_id=parent.id if parent else None,
        priority="中",
        status="待开始",
        progress=0,
        plan_start_date=seed_start,  # NOT NULL 约束,先占位,后由 _expand_parent_dates 收敛
        plan_end_date=seed_end,
        is_orphan=0
    )
    db.add(node)
    db.flush()
    cache[key] = node
    return node


def _expand_parent_dates(node, start, end):
    """把某父节点日期扩展为覆盖 [start, end](父色条应覆盖全部子任务)"""
    if start and (node.plan_start_date is None or start < node.plan_start_date):
        node.plan_start_date = start
    if end and (node.plan_end_date is None or end > node.plan_end_date):
        node.plan_end_date = end


@router.post("/project-plan/import")
async def import_tasks(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    从 xlsx 批量导入末级任务
    - 列结构见 EXPORT_COLUMNS(与导出一致)
    - 按 L1→L2→L3 路径 find-or-create 父链,末级(L4)挂靠到 L3 下
    - 责任人按 real_name、学校按 full_name 匹配;枚举非法/缺任务名 → 该行失败
    - 同路径+同名 L4 已存在 → 跳过(防重复导入);整批单事务,任一异常全回滚
    """
    from openpyxl import load_workbook

    fname = (file.filename or "").lower()
    if not fname.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="仅支持 .xlsx 文件")

    content = await file.read()
    if len(content) == 0:
        raise HTTPException(status_code=400, detail="文件为空")

    try:
        wb = load_workbook(BytesIO(content), data_only=True)
    except Exception:
        raise HTTPException(status_code=400, detail="无法解析 Excel 文件,请确认格式")
    ws = wb.active

    # 预建 名字->id 映射(责任人/学校)
    user_map = {u.real_name: u.id for u in db.query(User).all() if u.real_name}
    school_map = {s.full_name: s.id for s in db.query(School).all() if s.full_name}

    ok, skipped, failed = 0, 0, 0
    details = []
    parent_cache = {}
    year = str(date.today().year)

    try:
        rows = list(ws.iter_rows(min_row=2, values_only=True))
        for idx, row in enumerate(rows, start=2):
            row = list(row) + [""] * (len(EXPORT_COLUMNS) - len(row))
            l1, l2, l3, l4, assignee, school, d_start, d_end, st, pri, prog, note = row[:12]
            l1, l2, l3, l4 = _cell_str(l1), _cell_str(l2), _cell_str(l3), _cell_str(l4)
            assignee, school, st, pri = _cell_str(assignee), _cell_str(school), _cell_str(st), _cell_str(pri)
            note = _cell_str(note)

            # 整行空 → 跳过不计
            if not any([l1, l2, l3, l4, assignee, school]):
                continue

            # 必填校验
            if not l4:
                failed += 1
                details.append(f"第{idx}行: 任务名称(L4)为空")
                continue
            if not l1:
                failed += 1
                details.append(f"第{idx}行: L1阶段为空,无法确定层级归属")
                continue

            # 枚举校验(缺省给默认)
            st = st or "待开始"
            pri = pri or "中"
            if st not in _VALID_STATUS:
                failed += 1
                details.append(f"第{idx}行: 状态'{st}'非法(应为{'/'.join(_VALID_STATUS)})")
                continue
            if pri not in _VALID_PRIORITY:
                failed += 1
                details.append(f"第{idx}行: 优先级'{pri}'非法(应为高/中/低)")
                continue

            # 责任人/学校匹配
            assignee_id = user_map.get(assignee) if assignee else None
            if assignee and assignee_id is None:
                details.append(f"第{idx}行: 责任人'{assignee}'未匹配到用户,已留空")
            school_id_val = school_map.get(school) if school else None
            if school and school_id_val is None:
                details.append(f"第{idx}行: 学校'{school}'未匹配到,已留空")

            # 进度归一
            progress_val = _clamp_progress(prog) if prog not in (None, "") else 0

            # 日期必填校验(DDL: plan_start_date/plan_end_date NOT NULL)
            ps = _parse_date_cell(d_start)
            pe = _parse_date_cell(d_end)
            if ps is None or pe is None:
                failed += 1
                details.append(f"第{idx}行: 计划开始/结束日期缺失或格式非法(需 YYYY-MM-DD)")
                continue
            if pe < ps:
                failed += 1
                details.append(f"第{idx}行: 计划结束早于计划开始")
                continue

            # 重复校验:按完整路径签名 (L1/L2/L3/L4) 判重,不依赖 parent_id
            # (现有叶子跨级挂 L2,靠 parent_id 判重会永远落空)。
            # 先判重再建父节点,避免被跳过的行白建结构节点。
            # l3 历史数据可能是空串或 NULL,两者视为同一路径。
            def _eq_or_empty(col, val):
                if val == "":
                    return (col == "") | (col.is_(None))
                return col == val

            dup = db.query(WbsTask).filter(
                WbsTask.is_orphan == 0,
                WbsTask.project_phase_l1 == l1,
                _eq_or_empty(WbsTask.sub_phase_l2, l2),
                _eq_or_empty(WbsTask.task_package_l3, l3),
                WbsTask.work_content_l4 == l4
            ).first()
            if dup:
                skipped += 1
                details.append(f"第{idx}行: '{l4}'(路径 {l1}/{l2}/{l3})已存在,跳过")
                continue

            # 结构约定(与现有数据一致):结构节点只到 L2,L4 叶子直接挂 L2 下,
            # L3(任务包)作为叶子的内联字段存储,不物化成独立节点。
            # L2 为空时叶子挂 L1 下(兼容仅两级路径的行)。
            if l2:
                parent = _find_or_create_structure(db, 2, (l1, l2), year, parent_cache, ps, pe)
            else:
                parent = _find_or_create_structure(db, 1, (l1,), year, parent_cache, ps, pe)

            # 建 L4 叶子任务:L1/L2 取自父结构节点,L3 为内联字段,L4 为本行任务名
            new_task = WbsTask(
                task_code=_gen_task_code(db),
                construction_year=parent.construction_year or year,
                project_phase_l1=parent.project_phase_l1,
                sub_phase_l2=parent.sub_phase_l2,
                task_package_l3=l3,          # L3 内联存字段
                work_content_l4=l4,
                work_detail_l5="",
                parent_id=parent.id,
                priority=pri,
                status=st,
                progress=progress_val,
                plan_start_date=ps,
                plan_end_date=pe,
                responsible_person_id=assignee_id,
                school_id=school_id_val,
                progress_note=note
            )
            db.add(new_task)
            db.flush()

            # 扩展父结构节点(L1/L2)日期,使父色条覆盖全部子任务
            _expand_parent_dates(parent, ps, pe)
            l1_node = parent_cache.get((1, (l1,)))
            if l1_node:
                _expand_parent_dates(l1_node, ps, pe)

            ok += 1

        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500, detail=f"导入失败,已全部回滚: {e}")

    return {
        "success": ok,
        "skipped": skipped,
        "failed": failed,
        "total": ok + skipped + failed,
        "details": details[:100]  # 明细最多返回100条
    }
